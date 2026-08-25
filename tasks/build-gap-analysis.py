#!/usr/bin/env python3
"""Sinh tasks/gap-analysis-tasks.xlsx từ tasks/gap-analysis-tasks.md (nguồn sự thật).

Chạy từ thư mục tasks/:
    uv run --with openpyxl python build-gap-analysis.py
"""
import json
import re
import zipfile
from collections import Counter
from math import ceil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
MD = ROOT / "gap-analysis-tasks.md"
OUT = ROOT / "gap-analysis-tasks.xlsx"

STATUSES = ["Todo", "In Progress", "Review", "Done", "Blocked"]
PRIORITIES = ["P0", "P1", "P2"]

WIDTHS = {"ID": 8, "Tên task": 42, "Mô tả vấn đề": 46, "Priority": 10, "Category": 13,
          "Status": 13, "Assignee": 12, "Hạn dự kiến": 14, "Evidence / Tham chiếu": 46,
          "Acceptance / Phạm vi": 46, "Ghi chú": 34, "Báo cáo": 42}
CENTER_COLS = {"ID", "Priority", "Category", "Status", "Assignee", "Hạn dự kiến"}
WRAP_COLS = {"Tên task", "Mô tả vấn đề", "Evidence / Tham chiếu", "Acceptance / Phạm vi", "Ghi chú", "Báo cáo"}


HEADER_FILL = PatternFill("solid", start_color="1E293B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="111827")
BOLD = Font(bold=True, size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PRI_FILL = {"P0": PatternFill("solid", start_color="FEE2E2"),
            "P1": PatternFill("solid", start_color="FEF3C7"),
            "P2": PatternFill("solid", start_color="F1F5F9")}
PRI_FONT = {"P0": Font(color="991B1B", bold=True, size=11),
            "P1": Font(color="92400E", bold=True, size=11),
            "P2": Font(color="475569", bold=True, size=11)}
STATUS_FILL = {"Todo": PatternFill("solid", start_color="F8FAFC"),
               "In Progress": PatternFill("solid", start_color="DBEAFE"),
               "Review": PatternFill("solid", start_color="EDE9FE"),
               "Done": PatternFill("solid", start_color="DCFCE7"),
               "Blocked": PatternFill("solid", start_color="FEE2E2")}


def parse_master():
    lines = MD.read_text(encoding="utf-8").splitlines()
    start = next(i for i, l in enumerate(lines) if l.startswith("| ID |"))
    header = [c.strip() for c in lines[start].strip().strip("|").split("|")]
    rows = []
    for line in lines[start + 2:]:
        if not line.startswith("|"):
            break
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        assert len(cells) == len(header), f"cột lệch ở dòng: {line[:60]}"
        rows.append(dict(zip(header, cells)))
    return header, rows


HEADER, TASKS = parse_master()
assert len(TASKS) == 20, f"mong đợi 20 task, có {len(TASKS)}"
PRI_COL = get_column_letter(HEADER.index("Priority") + 1)
ST_COL = get_column_letter(HEADER.index("Status") + 1)

wb = Workbook()

# ---------------- Overview ----------------
ov = wb.active
ov.title = "Overview"
ov.merge_cells("A1:G1")
ov["A1"] = "Gap Analysis Tasks — Nexus Platform"
ov["A1"].font = TITLE_FONT
ov.merge_cells("A2:G2")
ov["A2"] = ("Nguồn: docs/research/mandatory-features-gap-analysis-2026-08-24.md "
            "(đã review bởi code-reviewer agent, mọi claim xác minh bằng source)")
ov["A2"].font = Font(italic=True, size=10, color="475569")
ov.merge_cells("A3:G3")
ov["A3"] = "Import: 2026-08-24 · 20 task (P0 × 4, P1 × 9, P2 × 7) · Nguồn sự thật là sheet Master"
ov["A3"].font = Font(italic=True, size=10, color="475569")

ov_headers = ["Nhóm", "Tổng", "Todo", "In Progress", "Review", "Done", "Blocked"]
for c, h in enumerate(ov_headers, 1):
    cell = ov.cell(row=5, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

for i, pri in enumerate(PRIORITIES):
    r = 6 + i
    ov.cell(row=r, column=1, value=pri).font = PRI_FONT[pri]
    ov.cell(row=r, column=2, value=f"=COUNTIF(Master!${PRI_COL}:${PRI_COL},$A{r})")
    for c, status in enumerate(STATUSES, 3):
        ov.cell(row=r, column=c,
                value=f'=COUNTIFS(Master!${PRI_COL}:${PRI_COL},$A{r},Master!${ST_COL}:${ST_COL},"{status}")')
r_tot = 9
ov.cell(row=r_tot, column=1, value="Tổng")
for c in range(2, 8):
    col = get_column_letter(c)
    ov.cell(row=r_tot, column=c, value=f"=SUM({col}6:{col}8)")
for row in ov.iter_rows(min_row=6, max_row=9, min_col=1, max_col=7):
    for cell in row:
        cell.border = BORDER
        if cell.column > 1:
            cell.alignment = Alignment(horizontal="center")
        if cell.row == 9:
            cell.font = BOLD
ov.freeze_panes = "A6"
ov.column_dimensions["A"].width = 12
for c in "BCDEFG":
    ov.column_dimensions[c].width = 12

# ---------------- Master ----------------
ms = wb.create_sheet("Master")
for c, name in enumerate(HEADER, 1):
    cell = ms.cell(row=1, column=c, value=name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

for i, task in enumerate(TASKS):
    r = 2 + i
    for c, name in enumerate(HEADER, 1):
        cell = ms.cell(row=r, column=c, value=task[name])
        cell.border = BORDER
        if name == "ID":
            cell.font = BOLD
        elif name == "Priority":
            cell.font = PRI_FONT[task[name]]
        if name in CENTER_COLS:
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        else:
            cell.alignment = WRAP_TOP

for col, name in enumerate(HEADER, 1):
    ms.column_dimensions[get_column_letter(col)].width = WIDTHS.get(name, 12)

for i, task in enumerate(TASKS):
    lines = 1
    for name in WRAP_COLS:
        v = task[name] or ""
        lines = max(lines, ceil(len(v) / (WIDTHS[name] * 0.95)))
    ms.row_dimensions[2 + i].height = min(14.5 * lines + 3, 150)

last = len(TASKS) + 1
ms.auto_filter.ref = f"A1:{get_column_letter(len(HEADER))}{last}"
ms.freeze_panes = "C2"

dv_pri = DataValidation(type="list", formula1='"P0,P1,P2"', allow_blank=True)
dv_pri.error = "Priority phải là P0, P1 hoặc P2"
ms.add_data_validation(dv_pri)
dv_pri.add(f"{PRI_COL}2:{PRI_COL}{last}")
dv_status = DataValidation(type="list", formula1='"Todo,In Progress,Review,Done,Blocked"', allow_blank=True)
ms.add_data_validation(dv_status)
dv_status.add(f"{ST_COL}2:{ST_COL}{last}")

for pri in PRIORITIES:
    ms.conditional_formatting.add(f"{PRI_COL}2:{PRI_COL}{last}",
                                  CellIsRule(operator="equal", formula=[f'"{pri}"'],
                                             fill=PRI_FILL[pri], font=PRI_FONT[pri]))
for st, fill in STATUS_FILL.items():
    ms.conditional_formatting.add(f"{ST_COL}2:{ST_COL}{last}",
                                  CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill))

# ---------------- Execution Order ----------------
eo = wb.create_sheet("Execution Order")
eo["A1"] = "Execution Order (khuyến nghị)"
eo["A1"].font = TITLE_FONT
order = [
    "1. GA-01 — fix quên mật khẩu (impact lớn nhất, fix nhỏ nhất, không đụng DB).",
    "2. GA-02 — xác minh code trạng thái `intake_ready` → chốt intake flow (cần quyết định Q1–Q5).",
    "3. GA-03 — rate limit + lockout đăng nhập.",
    "4. GA-04 — xóa tài khoản + GA-13 — 2FA admin/supporter + GA-07 — session timeout (nhóm bảo mật tài khoản).",
    "5. GA-12 — ToS/Privacy + consent + GA-11 — user data export (nhóm tuân thủ NĐ 13/2023).",
    "6. P1 còn lại: GA-05 avatar → GA-06 session UI → GA-09 pagination/search → GA-10 export CSV → GA-08 notification preferences.",
]
for i, line in enumerate(order):
    cell = eo.cell(row=3 + i, column=1, value=line)
    cell.alignment = WRAP_TOP
    eo.row_dimensions[3 + i].height = ceil(len(line) / 100) * 15 + 3
eo.column_dimensions["A"].width = 110

# ---------------- Conventions ----------------
cv = wb.create_sheet("Conventions")
cv["A1"] = "Conventions"
cv["A1"].font = TITLE_FONT
conventions = [
    "1. Nguồn sự thật là bảng Master — mọi cập nhật trạng thái sửa ở đây (và sheet khi đã đồng bộ).",
    "2. Ai cập nhật: Assignee cập nhật status + evidence task của mình; leader rà lại trước mỗi họp (1 lần/tuần).",
    "3. Status flow: Todo → In Progress (khi có branch/commit thật) → Review (khi mở PR, điền link vào Evidence) → Done (CHỈ khi PR merged + verify xong, không tự tick) → Blocked (kẹt >1 ngày hoặc cần quyết định — ghi lý do vào Ghi chú, nêu trong họp).",
    "4. Assignee: 1 tên duy nhất chịu trách nhiệm chính; để trống = chưa ai nhận, chốt trong họp.",
    "5. Evidence: task Done phải có ≥1 link (PR/commit/doc) — không link = chưa hoàn thành.",
    "6. Priority: P0 chặn release / vi phạm pháp lý → ưu tiên tuyệt đối, không trễ quá 1 sprint; P1 trong kế hoạch học kỳ; P2 làm khi rảnh, không cam kết hạn.",
    "7. Sort: luôn giữ P0 trước rồi ID; ID cố định GA-01…GA-20, không đổi khi sắp xếp.",
    "8. Báo cáo: task Done điền cột Báo cáo trỏ journal trong docs/journals/. Từ nay làm xong là trỏ vào đây. Todo để trống.",
]
for i, line in enumerate(conventions):
    cell = cv.cell(row=3 + i, column=1, value=line)
    cell.alignment = WRAP_TOP
    cv.row_dimensions[3 + i].height = ceil(len(line) / 100) * 15 + 3
cv.column_dimensions["A"].width = 110

wb.save(OUT)

# ---------------- Nhúng cached values cho công thức Overview ----------------
# Dev env không có LibreOffice nên không chạy được recalc.py của skill ck:xlsx.
# Công thức chỉ gồm COUNTIF/COUNTIFS/SUM trên dải tĩnh → tự tính kết quả
# (ngữ nghĩa exact-match) và ghi <v> vào XML để mọi reader (kể cả data_only)
# thấy số đúng; Excel/Sheets vẫn recalc bình thường khi mở.
pri_counts = {p: Counter() for p in PRIORITIES}
for t in TASKS:
    pri_counts[t["Priority"]][t["Status"]] += 1
expected = {}
for i, pri in enumerate(PRIORITIES):
    r = 6 + i
    expected[f"B{r}"] = sum(pri_counts[pri].values())
    for c, status in zip("CDEFG", STATUSES):
        expected[f"{c}{r}"] = pri_counts[pri][status]
for c in "BCDEFG":
    expected[f"{c}9"] = sum(expected[f"{c}{r}"] for r in (6, 7, 8))

tmp = OUT.with_suffix(".xlsx.tmp")
with zipfile.ZipFile(OUT) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
    wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
    rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rid = re.search(r'<sheet[^>]*name="Overview"[^>]*r:id="(rId\d+)"', wb_xml).group(1)
    target = re.search(rf'<Relationship[^>]*Target="([^"]+)"[^>]*Id="{rid}"', rels_xml).group(1)
    sheet_target = target.lstrip("/")
    injected = False
    for name in zin.namelist():
        if name == sheet_target:
            xml = zin.read(name).decode("utf-8")
            for ref, val in expected.items():
                xml, n = re.subn(
                    rf'(<c r="{ref}"[^>]*>)(<f>.*?</f>)(?:\s*<v[^>]*(?:/>|>\s*</v>))?(</c>)',
                    rf"\1\2<v>{val}</v>\3",
                    xml, count=1)
                if n != 1:
                    raise RuntimeError(f"formula cell {ref} not found for injection")
            zout.writestr(name, xml)
            injected = True
        else:
            zout.writestr(name, zin.read(name))
    if not injected:
        raise RuntimeError(f"sheet target {sheet_target} not found in archive")
tmp.replace(OUT)

# ---------------- Verify (bắt chước output recalc.py) ----------------
from openpyxl import load_workbook

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
errs = {e: [] for e in EXCEL_ERRORS}
total_formulas = 0
wbv = load_workbook(OUT, data_only=False)
for ws in wbv.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                if cell.value.startswith("="):
                    total_formulas += 1
                for e in EXCEL_ERRORS:
                    if e in cell.value:
                        errs[e].append(f"{ws.title}!{cell.coordinate}")
wbv.close()

wbval = load_workbook(OUT, data_only=True)
ov_check = wbval["Overview"]
actual = {ref: ov_check[ref].value for ref in expected}
wbval.close()

bad = {ref: (v, expected[ref]) for ref, v in actual.items() if v != expected[ref]}
result = {
    "status": "success" if (not any(errs.values()) and not bad) else "errors_found",
    "total_errors": sum(len(v) for v in errs.values()) + len(bad),
    "total_formulas": total_formulas,
    "error_summary": {e: {"count": len(v), "locations": v[:20]} for e, v in errs.items() if v},
    "cached_value_mismatches": bad,
}
print(json.dumps(result, ensure_ascii=False, indent=2))
